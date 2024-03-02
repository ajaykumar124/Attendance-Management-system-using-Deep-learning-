import glob
import face_recognition
import numpy as np
from datetime import datetime
import cv2
import os
import openpyxl
from tkinter import *
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk

# Initialize video capture
cap = cv2.VideoCapture(0)
FONT = cv2.FONT_HERSHEY_COMPLEX
images = []
names = []
is_processing = False  # Variable to track whether the system is processing

# Set up the correct path to the images directory
images_directory = r"D:\attendancesystem\images"
path = os.path.join(images_directory, "*.jpg")  # Assuming the images are in JPEG format

# Load images and names into lists
for file in glob.glob(path):
    image = cv2.imread(file)
    if image is not None:
        a = os.path.basename(file)
        b = os.path.splitext(a)[0]
        names.append(b)
        images.append(image)
    else:
        print(f"Failed to load image {file}")

def create_or_open_workbook():
    current_date = datetime.now().strftime("%Y-%m-%d")
    try:
        workbook = openpyxl.load_workbook(f"{current_date}.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.save(f"{current_date}.xlsx")
    sheet = workbook.active
    return workbook, sheet

def is_username_unique(sheet, username):
    for row in sheet.iter_rows(values_only=True):
        if row and row[0] == username:
            return False
    return True

def is_late_or_ontime(current_time_str, target_time_str, lateness_threshold_minutes):
    current_time = datetime.strptime(current_time_str, '%H:%M')
    target_time = datetime.strptime(target_time_str, '%H:%M')
    time_difference = current_time - target_time
    lateness_minutes = time_difference.total_seconds() / 60
    return "Late" if lateness_minutes >= lateness_threshold_minutes else "On Time"

def main(user_input, target_time_str):
    workbook, sheet = create_or_open_workbook()
    if sheet.max_row == 1 or sheet.max_row is None:
        sheet.append(["Username", "Current Date", "Current Time", "Mark"])
    if is_username_unique(sheet, user_input):
        current_time_str = datetime.now().strftime("%H:%M")
        lateness_threshold_minutes = 6
        result = is_late_or_ontime(current_time_str, target_time_str, lateness_threshold_minutes)
        current_date = datetime.now().strftime("%Y-%m-%d")
        current_time = datetime.now().strftime("%H:%M:%S")
        sheet.append([user_input, current_date, current_time, result])
        workbook.save(f"{current_date}.xlsx")

def encoding1(images):
    encode = []
    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        try:
            unk_encoding = face_recognition.face_encodings(img)[0]
            encode.append(unk_encoding)
        except IndexError:
            print(f"No face encodings found in image. Skipping image.")
    return encode

# Encode the loaded images
encodelist = encoding1(images)

# Tkinter window for input
def start_stop_attendance_system():
    global is_processing

    if not is_processing:
        target_time_str = target_time_entry.get()
        if target_time_str:
            try:
                datetime.strptime(target_time_str, '%H:%M')
            except ValueError:
                messagebox.showerror("Error", "Invalid time format. Please enter time in 'HH:MM' format (e.g., '13:45').")
                return

            is_processing = True
            start_btn.config(text="Stop", bg='#FF0000')
            target_time_entry.config(state=DISABLED)  # Disable the entry while processing

            def update_video_feed():
                ret, frame = cap.read()
                if ret:
                    frame_small = cv2.resize(frame, (0, 0), None, 0.25, 0.25)
                    face_locations = face_recognition.face_locations(frame_small)
                    curframe_encoding = face_recognition.face_encodings(frame_small, face_locations)

                    if encodelist:  # Check if encodelist is not empty
                        for encodeface, facelocation in zip(curframe_encoding, face_locations):
                            results = face_recognition.compare_faces(encodelist, encodeface)
                            distance = face_recognition.face_distance(encodelist, encodeface)
                            match_index = np.argmin(distance)

                            if results[match_index]:
                                name = names[match_index]
                                main(name, target_time_str)  # Pass the target time to the main function
                                y1, x2, y2, x1 = facelocation
                                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                                cv2.rectangle(frame, (y1, x1), (y2, x2), (0, 255, 0), 2)
                                cv2.putText(frame, name, (y1, x2 - 6), FONT, 1, (255, 0, 255), 2)
                            else:
                                y1, x2, y2, x1 = facelocation
                                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                                cv2.rectangle(frame, (y1, x1), (y2, x2), (0, 0, 255), 2)
                                cv2.putText(frame, "Unknown", (y1, x2 - 6), FONT, 1, (0, 0, 255), 2)

                        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                        img = Image.fromarray(cv2image)
                        imgtk = ImageTk.PhotoImage(image=img)
                        panel.imgtk = imgtk
                        panel.configure(image=imgtk)

                        if is_processing:
                            root.after(10, update_video_feed)
                    else:
                        messagebox.showerror("Error", "Encodelist is empty. Exiting program.")
                        root.destroy()
                else:
                    messagebox.showerror("Error", "Failed to grab frame. Exiting program.")
                    root.destroy()

            update_video_feed()

        else:
            messagebox.showerror("Error", "No target time provided. Exiting program.")

    else:
        is_processing = False
        start_btn.config(text="Start", bg='#008CBA')
        target_time_entry.config(state=NORMAL)  # Enable the entry after stopping

# Function to open the attendance file
def open_attendance_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        os.system(f'start excel "{file_path}"')

# GUI elements
root = Tk()
root.title("Attendance System")

# Load the background image
bg_image = Image.open("bg.jpg")
bg_photo = ImageTk.PhotoImage(bg_image)

# Configure the window size and position
root.geometry('%dx%d+0+0' % (root.winfo_screenwidth(), root.winfo_screenheight()))

# Background
background_label = Label(root, image=bg_photo)
background_label.place(relwidth=1, relheight=1)

# Placeholder for the video feed
panel = Label(root)
panel.pack()

# Target time entry
Label(root, text="Enter target time (HH:MM)").pack(pady=10)
target_time_entry = Entry(root, font=('calibre', 14, 'normal'))
target_time_entry.pack(pady=10)

# Start/Stop and Quit buttons
btn_frame = Frame(root)
btn_frame.pack(pady=20)
start_btn = Button(btn_frame, text="Start", command=start_stop_attendance_system, font=('calibre', 14, 'bold'), bg='#008CBA', fg='white')
start_btn.grid(row=0, column=0, padx=10)
quit_btn = Button(btn_frame, text="Quit", command=root.destroy, font=('calibre', 14, 'bold'), bg='#FF0000', fg='white')
quit_btn.grid(row=0, column=1, padx=10)

# Open Excel file button
open_file_btn = Button(root, text="Open Attendance File", command=open_attendance_file, font=('calibre', 14, 'bold'), bg='#4CAF50', fg='white')
open_file_btn.pack(pady=10)

# Label for displaying error messages
error_label = Label(root, text="", fg="red", font=('calibre', 12, 'bold'))
error_label.pack(pady=10)

root.mainloop()
